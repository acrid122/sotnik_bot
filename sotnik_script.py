import vk_api
import requests
import json
import time
from openpyxl import load_workbook
from vk_api.bot_longpoll import VkBotLongPoll
from vk_api.bot_longpoll import VkBotEventType
from vk_api.keyboard import VkKeyboard
from vk_api.keyboard import VkKeyboardColor
#from vk_api import VkUpload потом...
from pprint import pprint
from google_credits import getInfoAboutStudents
from google_credits import getInfoAboutHomeworks
from my_token import my_group_id
from my_token import my_token
from stepik_res import client_id_pr
from stepik_res import client_secret_pr
from stepik_credits import stepikAuthAndRetAmount
from datetime import datetime
'''from redis import Redis
from rq import Queue
q = Queue(connection=Redis())'''
students_amount = stepikAuthAndRetAmount()
ALL_USERS_IN_CURRENT_SESSION = []
students_otsr = {}
#Main Keyboard
def main_key_board():
   key_board = VkKeyboard(one_time = False, inline = False)
   key_board.add_button(
    label = "Я новенький!",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 1}
   )
   key_board.add_line()
   key_board.add_button(
    label = "Проблемы со Stepik",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 2}
   )
   key_board.add_line()
   key_board.add_button(
    label = "Как решать задачу",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 3}
   )
   key_board.add_line()
   key_board.add_button(
    label = "Жизни",
    color=VkKeyboardColor.POSITIVE,
    payload={'button': 4}
   )
   key_board.add_button(
    label = "Дедлайны",
    color=VkKeyboardColor.NEGATIVE,
    payload={'button': 5}
   )
   key_board.add_button(
    label = "Оплаты",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 6}
   )
   '''
   key_board.add_button(
    label = "Долги",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 7}
   )'''
   return key_board.get_keyboard()
#Dday Keyboard
def ddays_key_board():
   ddays_key_board = VkKeyboard(one_time = False, inline = False)
   ddays_key_board.add_button(
    label = "Узнать текущий дедлайн",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 15}
   )
   ddays_key_board.add_line()
   ddays_key_board.add_button(
    label = "Перенести дедлайн",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 16}
   )
   return ddays_key_board.get_keyboard()
#Ddline cancel reason
def dday_cancel_key_board():
   dday_cancel_key_board = VkKeyboard(one_time = False, inline = False)
   dday_cancel_key_board.add_button(
    label = "Тяжелая длительная болезнь",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 17}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Перегруз по учебе",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 18}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Отъезд",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 19}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Вынуждающие семейные обстоятельства",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 20}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Отсутствие доступа к интернету",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 21}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Другая причина",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 22}
   )
   dday_cancel_key_board.add_line()
   dday_cancel_key_board.add_button(
    label = "Вернуться в исходное меню",
    color=VkKeyboardColor.SECONDARY,
    payload={'button': 23}
   )
   return dday_cancel_key_board.get_keyboard()
#Chat Keyboard
def chats_key_board():
  chats_key_board = VkKeyboard(one_time = False, inline = False)
  chats_key_board.add_openlink_button(
    label = "Дежурный чат",
    link = 'https://vk.com/flashege23',
   )
  chats_key_board.add_line()
  chats_key_board.add_openlink_button(
    label = "Рабочий чат",
    link = 'https://vk.me/join/VjjdB9gMlZWfmXDpKjjndK4QxQBbHfBxpNg=',
   )
  chats_key_board.add_line()
  chats_key_board.add_openlink_button(
    label = "Флуд чат",
    link = 'https://vk.me/join/GFp8VUJ5TwGWGouG8uRrm7232LGm9vmiDpA='
   )
  chats_key_board.add_line()
  chats_key_board.add_button(
    label = "Вернуться в прошлое меню",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 8}
   )
  return chats_key_board.get_keyboard()
#Stepik key_board
def stepik_key_board():
  stepik_key_board = VkKeyboard(one_time = False, inline = False)
  stepik_key_board.add_button(
    label = "Не могу вступить в класс Stepik",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 9}
   )
  stepik_key_board.add_line()
  stepik_key_board.add_button(
    label = "Степик не начисляет баллы",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 10}
   )
  stepik_key_board.add_line()
  stepik_key_board.add_button(
    label = "Не открывается страница Stepik",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 11}
   )
  stepik_key_board.add_line()
  stepik_key_board.add_button(
    label = "Stepik не принимает решение",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 12}
   )
  stepik_key_board.add_line()
  stepik_key_board.add_button(
    label = "Не открываются решения на Stepik",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 13}
   )
  stepik_key_board.add_line()
  stepik_key_board.add_button(
    label = "Вернуться в исходное меню",
    color=VkKeyboardColor.SECONDARY,
    payload = {'button': 14}
   )
  return stepik_key_board.get_keyboard()
def check_cancel_about_studying(user_id):
   if students_otsr.get(user_id) == None:
      students_otsr[user_id] = 1
      return 0
   else:
      return 1
#Button Listener
def buttonKeyboardClick(user_id, payload, vk_session, own_vk_id, group_id):
  res = json.loads(payload)
  if 'button' in res:
    res = res['button']
    print(res)
    if res == 1:
      send_message(user_id, {'message': 'Доброго времени суток и добро пожаловать на курс!' +
                                ' Сейчас будет крайне важная информация, которую нужно прочитать' +
                                ' от начала до конца и задать вопросы, если они возникли.' + '\n\n'
                                '1. Сейчас у нас идет блок программирования, и мы' +
                                ' проходим его на платформе https://stepik.org . Для того, чтобы получить доступ к нашему' +
                                ' частному курсу по программированию, необходимо: зарегистрироваться,' +
                                ' скинуть сюда id и почту профиля (видео, как найти id есть в видеозаписях сообщества).' +
                                ' Доступ выдается в течение 1-2 дней' + '\n'
                                '2. Из-за позднего присоединения Вам придется очень стараться и много работать,' +
                                ' чтобы догнать материал, который на курсе уже пройден.' +
                                ' Я предлагаю такую схему: сдавать дз 3 раза в неделю, то есть будет 3 дедлайна:' +
                                ' во вторник, четверг и субботу. Дз можно сдавать в любой день, просто к' +
                                ' каждому из трех дней должна быть готова одна домашняя работа. Лучше всего будет' +
                                ' начать с блока по программированию, и выполнять все дз по порядку. После того,' +
                                ' как Вы догоните программирование, схема чуть-чуть поменяется: в четверг и субботу' +
                                ' нужно будет сдавать дз, как и все на курсе, то есть будут такие же дедлайны, как и у' +
                                ' всех, а ко вторнику нужно будет сдать любую из старых домашек уже с нашей платформы' +
                                ' 100points. Там я рекомендую выполнять дз по порядку.' +
                                ' Когда догоните программирование, нужно меня предупредить' + '\n' +
                                '3. Узнать баланс жизней, информацию об оплатах, перенести дедлайн' +
                                ' можно здесь, но есть строгие правила по переносу, которые описаны' +
                                ' в посте на стене этого сообщества. Любое общение со мной тоже будет происходит' +
                                ' здесь, не стоит писать мне в лс, там я на вопросы, связанные с сотниками, не отвечаю' +
                                ' Если остались вопросы, то нужно их задать тут, я обязательно на все отвечу!',
                                'dont_parse_links': 0}, vk_session)
    if res == 2:
      send_message(user_id, {'message': 'Выберите один из вопросов.',
                             'keyboard': stepik_key_board()}, vk_session)
    if res == 3:
      send_message(user_id, {'message': 'Данный диалог не предназначен'+
                             ' для вопросов по ДЗ.' + ' Если Вы хотите' +
                             ' узнать, как решать тот или иной номер - просьба' +
                             ' обратиться в дежурный чат (в личку сообщества). Или же задать вопрос' +
                             ' в одном из чатов.', 'keyboard': chats_key_board()}, vk_session)
    if res == 4:
      if lifeAmount(user_id) != 'Вас еще нет в группе':
        send_message(user_id, {'message': 'Количество твоих жизней: ' + lifeAmount(user_id)}, vk_session)
      else:
        send_message(user_id, {'message': 'Вас еще нет в группе.' +
                               ' Сообщение о необходимости добавить Вас на курс направлено сотнику.' }, vk_session)
        '''send_message(own_vk_id, {'message' : 'Необходимо добавить в журнал: vk_id = '
                                 + str(user_id)}, vk_session)'''
        vk_session.method('messages.send', {'user_id': own_vk_id,
                                            'random_id': 0,
                                            'message' : 'Необходимо добавить в журнал: https://vk.com/id' +
                                            str(user_id),
                                            'dont_parse_links': 0})
    if res == 5:
       send_message(user_id, {'message': 'Укажи, что именно Вы хотите сделать: ',
                             'keyboard': ddays_key_board()}, vk_session)
    if res == 6:
      checkOplati(user_id, group_id, vk_session)
      if checkOplati(user_id, group_id, vk_session) != 'Вас нет в списке оплат.':
        send_messge(user_id, {'message' : 'За этот месяц Вам надо заплатить: ' +
                              checkOplati(user_id, group_id, vk_session)}, vk_session)
      else:
        send_message(user_id, {'message': 'Вас еще нет в списке оплат или же список не обновлен.' +
                               ' Сообщение о необходимости разобрать со списком оплат направлено сотнику.' }, vk_session)
        vk_session.method('messages.send', {'user_id': own_vk_id,
                                            'random_id': 0,
                                            'message' : 'Необходимо посмотреть в списке оплат: https://vk.com/id' +
                                            str(user_id) + ' или просто посмотреть, что с самим списком.',
                                            'dont_parse_links': 0})
    '''if res == 7:
      checkHomeworks()'''
    if res == 8 or res == 14 or res == 23:
      send_message(user_id, {'message': 'Вы в исходном меню',
                             'keyboard' : main_key_board()}, vk_session)
    if res == 9:
      send_message(user_id, {'message': 'Скорее всего, у Вас нет доступа к курсу, когда его выдадут,' +
                             ' вступить в класс получиться. Если Вы еще не отправили данные,' +
                             ' для получения доступа, то нужно скорее это сделать.' +
                             ' Для получения доступы необходимы айди и почта'}, vk_session)
    if res == 10:
      send_message(user_id, {'message' : 'Это визуальный баг, нужно просто' +
                             'обновить страницу или немного подождать.'},
                   vk_session)
    if res == 11:
      send_message(user_id, {'message' : 'Платформа перегружена, нужно немного подождать.'},
                   vk_session)
    if res == 12:
      send_message(user_id, {'message' : 'Скорее всего, в решение было что-то не учтено.' +
                             ' Нужно еще немного подумать над ним.'},
                   vk_session)
    if res == 13:
      send_message(user_id, {'message' : 'Смотреть решения - это плохой подход к обучению,' +
                              ' попробуйте подумать сами или обратитесь к дежурному куратору.' +
                              ' Чтобы обратиться с вопросом в какой-либо чат, необходимо' +
                              ' перейти в исходное меню и нажать кнопку "Как решать задачу".'},
                   vk_session)
    if res == 15:
      send_message(user_id, {'message': 'Следующий дедлайн через: ' +
                             str(dDaysNext()[0]) +
                             ' Тема: ' + dDaysNext()[1]}, vk_session)
    if res == 16:
       send_message(user_id, {'message': 'Укажите причину переноса дедлайна. После указания причины, не забудьте' +
                              ' отправить справку или билет как подтверждение, начиная с фразы "ПД:подтверждение" ',
                              'keyboard' : dday_cancel_key_board()}, vk_session)
    if res == 17:
       send_message(user_id, {'message': 'Желаю Вам скорейшего выздоровления!' +
                              ' Сообщение передано сотнику.'}, vk_session)
       vk_session.method('messages.send', {'user_id': own_vk_id,
                                            'random_id': 0,
                                            'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                            str(user_id) + ' по причине тяжелой болезни.',
                                            'dont_parse_links': 0})
    if res == 18:
       if check_cancel_about_studying(user_id) == 0:
          send_message(user_id, {'message': 'Не забудьте, что предоставляется отсрочка по этой причине единоразово.' +
                                 ' Сообщение передано сотнику.'}, vk_session)
          vk_session.method('messages.send', {'user_id': own_vk_id,
                                               'random_id': 0,
                                               'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                               str(user_id) + ' по причине отсрочки по учебе.',
                                               'dont_parse_links': 0})
       else:
          send_message(user_id, {'message': 'Вам уже предоставлялась отсрочка по учебе.'+
                                 ' Сообщение передано сотнику.'}, vk_session)
          vk_session.method('messages.send', {'user_id': own_vk_id,
                                               'random_id': 0,
                                               'message' : 'Ученик: https://vk.com/id' +
                                               str(user_id) + ' уже переносил дедлайн по причине отсрочки по учебе.' +
                                               ' Нужно разобраться.',
                                               'dont_parse_links': 0})
    if res == 19:
       send_message(user_id, {'message': 'Приятной дороги Вам. Будь осторожен!' +
                              ' Сообщение передано сотнику.'}, vk_session)
       vk_session.method('messages.send', {'user_id': own_vk_id,
                                               'random_id': 0,
                                               'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                               str(user_id) + ' по причине отъезд.',
                                               'dont_parse_links': 0})
    if res == 20:
       send_message(user_id, {'message': 'Надеюсь, что все будет хорошо!'+
                              ' Сообщение передано сотнику.'}, vk_session)
       vk_session.method('messages.send', {'user_id': own_vk_id,
                                               'random_id': 0,
                                               'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                               str(user_id) + ' по семейным обстоятельствам.',
                                               'dont_parse_links': 0})
    if res == 21:
       send_message(user_id, {'message': 'Надеюсь, скоро Вас подключат к сети...'+
                              ' Сообщение передано сотнику.'}, vk_session)
       vk_session.method('messages.send', {'user_id': own_vk_id,
                                               'random_id': 0,
                                               'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                               str(user_id) + ' по причине отсутствия Интернета.',
                                               'dont_parse_links': 0})
    if res == 22:
       send_message(user_id, {'message': 'Введите Вашу причину, начиная с фразы "ПД:Другая причина" '}, vk_session)
''' Потом def vkUploadPhoto():
   vk_upload = VkUpload()
   return vk_upload'''
#Vk Send message
def send_message(user_id, message, vk_session):
  message_payload = {'user_id': user_id, 'random_id': 0, **message}
  vk_session.method('messages.send', message_payload)
#Vk LongPoll Api
def vkLongPoll(vk_session, group_id):
  longpoll = VkBotLongPoll(vk_session, group_id)
  for event in longpoll.listen():
    '''cur_time = datetime.now()
    q.enqueue_at(cur_time, getInfoAbountStudents)'''
    if event.type == VkBotEventType.MESSAGE_NEW:
      user_message = event.object.message['text'].lower()
      user_id = event.object.message['from_id']
      if user_id not in ALL_USERS_IN_CURRENT_SESSION:
        ALL_USERS_IN_CURRENT_SESSION.append(user_id)
        send_message(user_id, {'message': 'Бот "Сотник" приветствует Вас!', 'keyboard': main_key_board()}, vk_session)
      if user_message.find("пд:другая причина") != -1:
         send_message(user_id, {'message': 'Сообщение было передано куратору.'},vk_session)
         vk_session.method('messages.send', {'user_id': findOwnVkId(),
                                               'random_id': 0,
                                               'message' : 'Необходимо перенести дедлайн у: https://vk.com/id' +
                                               str(user_id) + ' по причине' + '\n' + user_message,
                                               'dont_parse_links': 0})
      if 'payload' in event.object.message:
        buttonKeyboardClick(user_id, event.object.message['payload'], vk_session, findOwnVkId(), group_id)
      if len(event.object.message['attachments']) > 0 and \
         event.object.message['attachments'][0].get('type') == 'photo' and \
         user_message.find("пд:подтверждение") != -1:
            vk_session.method('messages.send', {'user_id': findOwnVkId(),
                                                  'random_id': 0,
                                                  'message' : 'Отправил фотку с подтверждением отсрочки: https://vk.com/id' +
                                                  str(user_id) + ' в ' + str(datetime.now())})
      elif user_message.find("пд:подтверждение") != -1:
         send_message(user_id, {'message': 'Вы не прикрепили фотку.'},vk_session)
#Start Vk Session
def vkBotSession():
  vk_session = vk_api.VkApi(token = my_token)
  return vk_session
#Own Vk_id
def findOwnVkId():
  own_vk_id = vkBotSession().method('utils.resolveScreenName', {'screen_name': 'suyuki_cream'})['object_id']
  return own_vk_id
#Check Homework
'''def checkHomeworks():
  all_homework_list = getInfoAboutHomeworks()['values']
  homeworks_name = getInfoAboutHomeworks()['values'][0]
  max_homework_grade = getInfoAboutHomeworks()['values'][2]
  all_students_grades = [0]*(students_amount + 2)
  homeworks_name_and_grade = {}
  homeworks_name_and_cur = {}
  #Двумерный массив учеников
  for i in range(0, students_amount):
    all_students_grades[i] = getInfoAboutHomeworks()['values'][i + 6]
    #Добивание длины массива до числа домашек
    if(len(all_students_grades[i]) < len(homeworks_name)):
      for k in range(len(all_students_grades[i]), len(homeworks_name)):
        all_students_grades.append('')
  for i in range(0, len(homeworks_name)):
    print(homeworks_name[i])
    if homeworks_name[i] != '':
      #Формирование словаря 'название_домашки' : макс_балл
      homeworks_name_and_grade[homeworks_name[i]] = max_homework_grade[i]
    for k in range(0, len(all_students_grades)):
      print(all_students_grades[k][i])
        #Формирование словаря 'название_домашки' : лист из оценок учеников
      homeworks_name_and_cur[homeworks_name[i]] = str(all_students_grades[k][i])
  pprint(homeworks_name_and_cur)'''
#Get Group Members
def get_members(group_id, vk_session):
  group_getMembers = {'group_id': group_id, 'offset': 0}
  all_members = vk_session.method('groups.getMembers', group_getMembers)['items']
  return all_members
#Check Oplati
def checkOplati(user_id, group_id, vk_session):
  flash_wb = load_workbook('./Оплаты.xlsx')
  flash_gr = flash_wb.get_sheet_by_name('Лист1')
  all_members = get_members(group_id, vk_session)
  for i in range(1, len(all_members)):
    tmp = flash_gr.cell(row = i, column = 3).value
    if tmp is not None:
      if int(tmp[tmp.find("id")+2:]) in all_members:
        return str(flash_gr.cell(row = i, column = 9).value)
  return 'Тебя нет в списке оплат.'
#Check Life Amount
def lifeAmount(user_id):
  lifes_list = getInfoAboutStudents()['values']
  fl = False
  for line in lifes_list:
    if len(line) != 0 and line[2][line[2].find("id")+2:] == str(user_id):
      print(line[2])
      fl = True
      return line[4]
  if fl == False:
    return 'Тебя еще нет в группе'
#Make Date of DD 
def dDaysTimeChange(change_time):
  if change_time.find('day') != -1:
    if change_time.find('day') + 3 == 's':
      change_time = change_time[:change_time.find('day') + 5]
    else:
      change_time = change_time[:change_time.find('day') + 4]
    print(change_time)
  else:
    change_time = "0 days. Дедлайн сегодня!"
  return change_time
#Check DD
def dDaysNext():
   f = open('ddays.txt', encoding='utf-8')
   cur_time = datetime.now()
   diff_time = {}
   info_dday_time = {}
   for line in f:
     pr = []
     for i in range(0, len(line)):
       if line[i] == ' ':
         pr.append(i)
     year = line[:pr[0]]
     month = line[pr[0] + 1: pr[1]]
     day = line[pr[1] + 1: pr[2]]
     info = line[pr[2] + 1:]
     dday_time = datetime(int(year), int(month), int(day),23,59,59)
     if dday_time - cur_time >= cur_time - cur_time:
        diff_time[dday_time] = dday_time - cur_time
        info_dday_time[dday_time] = info
   diff_time = sorted(diff_time.items())
   info_dday_time = sorted(info_dday_time.items())
   info_dday = info_dday_time[0]
   info_dday = info_dday[1][:len(info_dday[1])-1]
   needed_item_time = diff_time[0]
   delta_needed_item_time = needed_item_time[1]
   delta_needed_item_time = dDaysTimeChange(str(delta_needed_item_time))
   f.close()
   return [delta_needed_item_time, info_dday]
if __name__ == "__main__":
  group_id = my_group_id
  vkLongPoll(vkBotSession(), group_id)
